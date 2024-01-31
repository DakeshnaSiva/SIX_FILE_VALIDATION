import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

def process_addition(data, file2):
    mc = data["Main_column"]
    rv = data["Column_RV"]
    ch = str(data["What to check?"]).split('\n')  # Ensure it's a string

    status_list = []
    expected_list = []
    actual_list = []

    # Check if Column_RV value is not found initially
    if file2[file2[mc] == rv].empty:
        status_list.append(f"{mc}={rv} not found")
        return '', '', '\n'.join(status_list)

    for check in ch:
        if check and '=' in check:
            check_column, check_value = check.split('=')
            check_value = float(check_value)

            # Check if the specified column exists in file2
            if check_column not in file2.columns:
                status_list.append(f'{check_column} column not found in file2')
                continue

            actual_values = file2.loc[file2[mc] == rv, check_column].tolist()

            if actual_values:
                expected_value = 0  # Change expected value to 0
                actual_value = actual_values[0]
                
                # Check if the actual value is numeric, replace with 0 if not
                actual_value = pd.to_numeric(actual_value, errors='coerce')
                actual_value = 0 if pd.isna(actual_value) else actual_value
                
                status = 'Success' if expected_value == actual_value else 'Fail'
                expected_list.append(f"{check_column}={expected_value}")
                actual_list.append(f"{check_column}={actual_value}")
            else:
                status = f'{check_column}={check_value} value not found'

            status_list.append(status)

    return '\n'.join(map(str, expected_list)), '\n'.join(map(str, actual_list)), '\n'.join(status_list)


def highlight_fail_values(sheet, expected_values_column, actual_column):
    red_fill = Font(color='FFFF0000')  # Red color

    # Iterate through the cells in the 'Expected' column to find the matching column
    for col in sheet.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value == expected_values_column:
                expected_column_index = cell.column

                # Iterate through the rows, starting from the second row
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=expected_column_index, max_col=expected_column_index):
                    expected_cell = row[0]
                    actual_cell = expected_cell.offset(column=1)

                    # Check if both cells have values
                    if expected_cell.value is not None and actual_cell.value is not None:
                        expected_values = expected_cell.value.split('\n')
                        actual_values = actual_cell.value.split('\n')

                        # Iterate through the values in the cells
                        for i in range(len(actual_values)):
                            if actual_values[i] != expected_values[i]:
                                # Highlight only the mismatched part of the cell value
                                actual_cell.value = actual_cell.value.replace(actual_values[i], f"<fill>{actual_values[i]}<fill>")
                                actual_cell.offset(row=i).font = red_fill

                                # Remove <fill> from the cell value
                                actual_cell.value = actual_cell.value.replace('<fill>', '')


def main():
    # Update file paths for the client system
    file1 = pd.read_excel(r'D:\dakeshna\SIX_File Validation\Book1.xlsx')
    file2 = pd.read_csv(r'D:\dakeshna\SIX_File Validation\check.csv',delimiter=";")

    file1['Status'] = ''
    file1['Expected'] = ''
    file1['Actual'] = ''

    for row, data in file1.iterrows():
        expected, actual, status = process_addition(data, file2)
        file1.at[row, 'Expected'] = expected
        file1.at[row, 'Actual'] = actual
        file1.at[row, 'Status'] = status


    file1['Actual'] = file1['Actual'].replace('nan', '0')

    file1 = file1.drop(columns=['Type'])

    file1.to_excel(r'D:\dakeshna\SIX_File Validation\File_1_updated.xlsx', index=False)

    wb = load_workbook(r'D:\dakeshna\SIX_File Validation\File_1_updated.xlsx')
    sheet = wb.active

    highlight_fail_values(sheet, 'Expected', 'Actual')
    wb.save(r'D:\dakeshna\SIX_File Validation\File_1_updated_highlighted_values.xlsx')


if __name__ == "__main__":
    main()